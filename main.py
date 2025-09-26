TEST_ONLY = False  # 设置为 True 时只进行评估，不进行训练

import random
import numpy as np
import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
from collections import deque
from tqdm import tqdm
import os
import matplotlib.pyplot as plt  # 确保已导入matplotlib
from datasets import ImbalancedDataset
from Model import Q_Net_image, TBM_conv1d, TBM_conv1d_1layer, TBM_conv1d_3layer, TBM_conv1d_4layer, TBM_conv1d_5layer, TBM_conv1d_6layer, TBM_conv1d_7layer, TBM_conv1d_8layer, ResNet32_1D, LSTM, BiLSTM, Transformer  # 导入所有模型类
from evaluate import evaluate_model  # 导入评估模块
import pandas as pd

def set_random_seed(seed):
    """
    设置所有随机数种子以确保实验的可重现性
    
    Args:
        seed: 随机数种子
    """
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    # 确保CUDA操作的确定性
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False
    print(f"已设置随机数种子: {seed}")

#由于方差太小，此次数据用标准差
def calculate_and_update_variance(save_dir, dataset_name, training_ratio, num_runs, rho):
    """
    计算最近num_runs次训练的各类准确率标准差并更新Excel文件
    
    Args:
        save_dir: 保存目录
        dataset_name: 数据集名称
        training_ratio: 训练完成比例
        num_runs: 运行次数
        rho: 不平衡率
    """
    excel_path = os.path.join(save_dir, 'evaluation_results.xlsx')
    
    if not os.path.exists(excel_path):
        print(f"错误: 未找到Excel文件 {excel_path}")
        return
    
    try:
        # 读取现有数据
        df = pd.read_excel(excel_path, header=0)
        
        # 筛选出最近num_runs次的相同数据集、训练比例和不平衡率的记录
        filtered_df = df[
            (df['数据集名称'] == dataset_name) & 
            (df['训练完成比例'] == training_ratio) &
            (df['不平衡率rho'] == rho)
        ].tail(num_runs)
        
        if len(filtered_df) < num_runs:
            print(f"警告: 只找到 {len(filtered_df)} 条记录，少于期望的 {num_runs} 次")
        
        # 要计算标准差的指标列表
        metrics = ['总体准确率', '头部准确率', '中部准确率', '尾部准确率']
        
        # 计算并添加每个指标的标准差
        for metric in metrics:
            # 提取指标值
            metric_values = filtered_df[metric].values
            
            # 计算标准差
            metric_std = np.std(metric_values, ddof=1)  # 使用样本标准差
            
            print(f"{metric}值: {metric_values}")
            print(f"{metric}标准差: {metric_std:.6f}")
            
            # 添加标准差列（如果不存在）
            std_col_name = f'{metric}标准差'
            if std_col_name not in df.columns:
                df[std_col_name] = None
            
            # 获取最近num_runs次记录的索引
            recent_indices = filtered_df.index
            
            # 将标准差值添加到这些行
            for idx in recent_indices:
                df.loc[idx, std_col_name] = metric_std
        
        # 保存更新后的Excel文件
        df.to_excel(excel_path, index=False, header=True)
        print(f"各指标标准差已添加到Excel文件: {excel_path}")
        
        # 使用openpyxl进行单元格合并
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment
            
            # 加载工作簿
            wb = load_workbook(excel_path)
            ws = wb.active
            
            # 对每个标准差列进行单元格合并
            for metric in metrics:
                std_col_name = f'{metric}标准差'
                
                # 找到标准差列的位置
                std_col = None
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=1, column=col).value == std_col_name:
                        std_col = col
                        break
                
                if std_col:
                    # 找到需要合并的行范围（最近num_runs次记录）
                    start_row = recent_indices[0] + 2  # +2 因为Excel从1开始，且有标题行
                    end_row = recent_indices[-1] + 2
                    
                    # 合并单元格
                    if len(recent_indices) > 1:
                        merge_range = f"{ws.cell(row=start_row, column=std_col).coordinate}:{ws.cell(row=end_row, column=std_col).coordinate}"
                        ws.merge_cells(merge_range)
                        
                        # 设置居中对齐
                        merged_cell = ws.cell(row=start_row, column=std_col)
                        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                        merged_cell.value = df.loc[recent_indices[0], std_col_name]
                        
                        print(f"已合并单元格 {merge_range} 并设置{std_col_name}值")
            
            # 保存工作簿
            wb.save(excel_path)
            print("Excel文件更新完成，单元格已合并")
            
        except ImportError:
            print("警告: 未安装openpyxl，无法进行单元格合并")
        except Exception as e:
            print(f"单元格合并时出错: {e}")
    
    except Exception as e:
        print(f"处理Excel文件时出错: {e}")

class MyRL():
    def __init__(self, input_shape, rho=0.01, model_type='Q_Net_image', reward_multiplier=1.0, discount_factor=0.1, num_classes=9):

        self.discount_factor = discount_factor  # 使用传入的折扣因子参数
        self.mem_size = 50000
        self.rho = rho
        self.reward_multiplier = reward_multiplier  # 保存奖励倍数
        self.num_classes = num_classes  # 类别数量（默认为9个类别）
        self.t_max = 120000
        self.eta = 0.05
        self.learning_rate = 0.00025
        self.batch_size = 64
        self.ratio = 1
        
        # 添加用于记录损失的列表
        self.loss_history = []
        
        # 根据模型类型选择网络，输出维度改为num_classes
        if model_type == 'Q_Net_image':
            self.q_net = Q_Net_image(input_shape, output_dim=num_classes)
            self.target_net = Q_Net_image(input_shape, output_dim=num_classes)
        elif model_type == 'TBM_conv1d':
            self.q_net = TBM_conv1d(input_shape, output_dim=num_classes)
            self.target_net = TBM_conv1d(input_shape, output_dim=num_classes)
        elif model_type == 'TBM_conv1d_1layer':
            self.q_net = TBM_conv1d_1layer(input_shape, output_dim=num_classes)
            self.target_net = TBM_conv1d_1layer(input_shape, output_dim=num_classes)
        elif model_type == 'TBM_conv1d_3layer':
            self.q_net = TBM_conv1d_3layer(input_shape, output_dim=num_classes)
            self.target_net = TBM_conv1d_3layer(input_shape, output_dim=num_classes)
        elif model_type == 'TBM_conv1d_4layer':
            self.q_net = TBM_conv1d_4layer(input_shape, output_dim=num_classes)
            self.target_net = TBM_conv1d_4layer(input_shape, output_dim=num_classes)
        elif model_type == 'TBM_conv1d_5layer':
            self.q_net = TBM_conv1d_5layer(input_shape, output_dim=num_classes)
            self.target_net = TBM_conv1d_5layer(input_shape, output_dim=num_classes)
        elif model_type == 'TBM_conv1d_6layer':
            self.q_net = TBM_conv1d_6layer(input_shape, output_dim=num_classes)
            self.target_net = TBM_conv1d_6layer(input_shape, output_dim=num_classes)
        elif model_type == 'TBM_conv1d_7layer':
            self.q_net = TBM_conv1d_7layer(input_shape, output_dim=num_classes)
            self.target_net = TBM_conv1d_7layer(input_shape, output_dim=num_classes)
        elif model_type == 'TBM_conv1d_8layer':
            self.q_net = TBM_conv1d_8layer(input_shape, output_dim=num_classes)
            self.target_net = TBM_conv1d_8layer(input_shape, output_dim=num_classes)
        elif model_type == 'ResNet32_1D':
            self.q_net = ResNet32_1D(input_shape, output_dim=num_classes)
            self.target_net = ResNet32_1D(input_shape, output_dim=num_classes)
        elif model_type == 'LSTM':
            self.q_net = LSTM(input_shape, output_dim=num_classes)
            self.target_net = LSTM(input_shape, output_dim=num_classes)
        elif model_type == 'BiLSTM':
            self.q_net = BiLSTM(input_shape, output_dim=num_classes)
            self.target_net = BiLSTM(input_shape, output_dim=num_classes)
        elif model_type == 'Transformer':
            self.q_net = Transformer(input_shape, output_dim=num_classes)
            self.target_net = Transformer(input_shape, output_dim=num_classes)
        else:
            raise ValueError(f"不支持的模型类型: {model_type}")
        
        self.target_net.load_state_dict(self.q_net.state_dict())  # 同步参数

        # 优化器
        self.optimizer = optim.Adam(self.q_net.parameters(), lr=self.learning_rate)

        # 经验回放池
        self.replay_memory = deque(maxlen=self.mem_size)

        # 设备配置
        self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        
        # 将网络移动到设备
        self.q_net.to(self.device)
        self.target_net.to(self.device)

        # 训练计数器
        self.step_count = 0
        self.epsilon = 1.0
        self.epsilon_min = 0.01
        self.epsilon_decay = (self.epsilon - self.epsilon_min) / (self.t_max*self.ratio)

        # 存储模型类型，用于后续数据预处理
        self.model_type = model_type
        
        # 奖励权重字典 - 将从数据集中获取
        self.reward_weights = {}

    def set_reward_weights(self, reward_weights):
        """设置各类别的奖励权重"""
        self.reward_weights = reward_weights
        print(f"已设置奖励权重: {self.reward_weights}")

    def compute_reward(self, action, label):
        """
        多分类问题的奖励函数
        Args:
            action: 预测的类别 (0-8)
            label: 真实的类别 (0-8)
        Returns:
            reward: 奖励值
            terminal: 是否终止当前episode
        """
        terminal = False
        
        # 获取当前类别的奖励权重（如果不存在则默认为1.0）
        weight = self.reward_weights.get(label, 1.0)
        
        # 预测正确的情况
        if action == label:
            reward = weight * self.reward_multiplier  # 正确分类奖励，乘以权重和倍率
        # 预测错误的情况
        else:
            reward = -weight * self.reward_multiplier  # 错误分类惩罚，乘以权重和倍率
            # 如果错分类别8（最少样本），则终止episode
            if label == 8:
                terminal = True
                
        return reward, terminal

    def replay_experience(self, update_target=True):
        """从经验回放缓冲区采样并训练网络"""                
        # 随机采样一批经验
        batch = random.sample(self.replay_memory, self.batch_size)
        states, actions, rewards, next_states, terminals = zip(*batch)

        # 将数据移动到正确的设备
        states = torch.stack(states).to(self.device)
        actions = torch.tensor(actions, dtype=torch.int64, device=self.device).unsqueeze(1)
        rewards = torch.tensor(rewards, dtype=torch.float32, device=self.device).unsqueeze(1)
        next_states = torch.stack(next_states).to(self.device)
        terminals = torch.tensor(terminals, dtype=torch.bool, device=self.device).unsqueeze(1)
        
        # 对于TBM模型，确保数据维度正确
        if self.model_type == 'TBM_conv1d':
            # TBM模型期望输入形状为[batch, channels, length]
            # 而我们的数据形状为[batch, length, channels]
            # 所以需要转置为[batch, channels, length]
            if states.shape[1] != 3 and states.shape[2] == 3:  # 如果是[batch, length, channels]
                states = states.transpose(1, 2)  # 转为[batch, channels, length]
                next_states = next_states.transpose(1, 2)

        # 计算当前Q值
        current_q = self.q_net(states).gather(1, actions)
            
        # 计算目标Q值
        with torch.no_grad():
           next_q = self.target_net(next_states).max(1, keepdim=True)[0]
           target_q = rewards + self.discount_factor * next_q * (~terminals)
            
        # 计算损失并更新
        loss = F.mse_loss(current_q, target_q)
        
        # 记录损失值
        self.loss_history.append(loss.item())

        # 清零梯度，反向传播，更新参数
        self.optimizer.zero_grad()
        loss.backward()
        self.optimizer.step()
            
        # 更新目标网络 (软更新)，只在update_target为True时更新
        # 更新参数 φ := (1-η)φ + ηθ
        if update_target:
            for target_param, param in zip(self.target_net.parameters(), self.q_net.parameters()):
                target_param.data.copy_(self.eta * param.data + (1.0 - self.eta) * target_param.data)
            
        # 衰减探索率
        if self.epsilon > self.epsilon_min:
            self.epsilon -= self.epsilon_decay
            

    def train(self, dataset):
        """
        按照论文Algorithm 2训练DQN分类器
        Args:
            dataset: 数据集对象
        """
        # 获取完整数据集
        train_data, train_labels, _, _ = dataset.get_full_dataset()
        
        # 从数据集获取奖励权重
        dist_info = dataset.get_class_distribution()
        if dist_info["reward_weights"] is not None:
            self.set_reward_weights(dist_info["reward_weights"])
            
        self.step_count = 0
        episode = 0
        
        # 创建总体训练进度条
        total_pbar = tqdm(total=self.t_max, desc="Training Progress", unit="step")
        
        # 外层循环: for episode k = 1 to K do (直到达到最大步数)
        while self.step_count < self.t_max:
            episode += 1
            
            # 打乱训练数据顺序 (Shuffle the training data D)
            indices = torch.randperm(len(train_data))
            shuffled_data = train_data[indices]
            shuffled_labels = train_labels[indices]
            
            # 初始化状态 s_1 = x_1
            current_state = shuffled_data[0:1]
            
            # 根据模型类型进行不同的数据预处理
            if self.model_type == 'TBM_conv1d':
                # TBM数据不需要添加通道维度，直接使用原始形状
                current_state = current_state.float().to(self.device)
            else:
                # 图像数据需要添加通道维度
                if len(current_state.shape) == 3:
                    current_state = current_state.unsqueeze(1)  # 添加通道维度
                current_state = current_state.float().to(self.device)
                
                # 修正通道顺序
                if current_state.shape[1] != 3 and current_state.shape[-1] == 3:
                    current_state = current_state.permute(0, 3, 1, 2)  # NHWC -> NCHW
            
            # 进度条显示
            episode_pbar = tqdm(
                total=len(shuffled_data)-1,  # 最后一个样本没有next_state
                desc=f"Episode {episode}", 
                leave=False, 
                unit="sample"
            )
            
            # 内层循环: for t = 1 to T do (遍历所有样本)
            for t in range(len(shuffled_data) - 1):
                # 检查是否已达到最大步数
                if self.step_count >= self.t_max:
                    break
                
                # 获取当前标签
                current_label = shuffled_labels[t].item()
                
                # 根据ε-greedy策略选择动作 (Choose an action based on ε-greedy policy)
                if random.random() < self.epsilon:
                    action = random.randint(0, self.num_classes - 1)  # 随机探索，范围0-8
                else:
                    with torch.no_grad():
                        q_values = self.q_net(current_state)
                    action = q_values.argmax().item()
                
                # 计算奖励和终止标志 (r_t, terminal_t = STEP(a_t, l_t))
                reward, terminal = self.compute_reward(action, current_label)
                
                # 获取下一状态 (Set s_{t+1} = x_{t+1})
                next_state = shuffled_data[t+1:t+2]
                
                # 根据模型类型进行不同的数据预处理
                if self.model_type == 'TBM_conv1d':
                    # TBM数据不需要添加通道维度
                    next_state = next_state.float().to(self.device)
                else:
                    # 图像数据需要添加通道维度
                    if len(next_state.shape) == 3:
                        next_state = next_state.unsqueeze(1)
                    next_state = next_state.float().to(self.device)
                    
                    # 修正通道顺序
                    if next_state.shape[1] != 3 and next_state.shape[-1] == 3:
                        next_state = next_state.permute(0, 3, 1, 2)
                
                # 存储经验到记忆库 (Store (s_t, a_t, r_t, s_{t+1}, terminal_t) to M)
                self.replay_memory.append((
                    # 直接存储原始形状，不进行维度转换
                    current_state.squeeze(0).cpu().clone().detach(),
                    action,
                    reward,
                    next_state.squeeze(0).cpu().clone().detach(),
                    terminal
                ))
                
                # 从记忆库中采样并学习(仅当记忆库足够大时)
                if len(self.replay_memory) >= self.batch_size:
                    # 根据terminal状态决定是否更新目标网络
                    self.replay_experience(update_target=not terminal)

                    self.step_count += 1
                    total_pbar.update(1)
                
                # 更新进度条
                episode_pbar.update(1)
                episode_pbar.set_postfix({
                    'Step': self.step_count,
                    'Epsilon': f'{self.epsilon:.4f}',
                    'Reward': f'{reward:.4f}',
                    'Terminal': terminal
                })
                
                # 如果是terminal状态，则终止当前episode
                if terminal:
                    break
                    
                # 设置当前状态为下一状态，继续循环
                current_state = next_state
            
            episode_pbar.close()
            
            # 显示episode信息
            total_pbar.set_postfix({
                'Episode': episode,
                'Epsilon': f'{self.epsilon:.4f}',
                'Memory': len(self.replay_memory)
            })
        
        total_pbar.close()
        print("训练完成!")
    
    
    def plot_loss(self, save_path):
        """
        绘制训练过程中的损失曲线
        Args:
            save_path: 保存图片的路径
        """
        plt.figure(figsize=(10, 6))
        plt.plot(self.loss_history)
        #plt.title('Q-Network Training Loss')
        plt.xlabel('Training Steps', fontsize=18)
        plt.ylabel('Loss', fontsize=18)
        plt.xticks(fontsize=18)
        plt.yticks(fontsize=18)
        plt.grid(False)
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        plt.close()
        print(f"损失曲线已保存到 {save_path}")

def get_model_config(dataset_name, model_variant=None):
    """
    根据数据集名称和模型变体返回相应的模型配置
    
    Args:
        dataset_name: 数据集名称
        model_variant: 模型变体，例如 'TBM_conv1d_1layer'
        
    Returns:
        dict: 包含模型类型和输入形状的配置
    """
    # 图像数据集配置
    image_datasets = {
        'mnist': {'model_type': 'Q_Net_image', 'input_shape': (1, 28, 28)},
        'fashion_mnist': {'model_type': 'Q_Net_image', 'input_shape': (1, 28, 28)},
        'cifar10': {'model_type': 'Q_Net_image', 'input_shape': (3, 32, 32)},
        'cifar100': {'model_type': 'Q_Net_image', 'input_shape': (3, 32, 32)},
    }
    
    # TBM数据集配置 - 使用1D卷积模型
    tbm_datasets = {
        'TBM_K': {'model_type': 'TBM_conv1d', 'input_shape': (1024, 3)},  # (len_window, feature_dim)
        'TBM_M': {'model_type': 'TBM_conv1d', 'input_shape': (1024, 3)},
        'TBM_K_M': {'model_type': 'TBM_conv1d', 'input_shape': (1024, 3)},
        'TBM_K_Noise': {'model_type': 'TBM_conv1d', 'input_shape': (1024, 3)},
        'TBM_M_Noise': {'model_type': 'TBM_conv1d', 'input_shape': (1024, 3)},
        'TBM_K_M_Noise': {'model_type': 'TBM_conv1d', 'input_shape': (1024, 3)},
        # 添加新的数据集配置
        'TBM_K_M_Noise_snr_3': {'model_type': 'TBM_conv1d', 'input_shape': (1024, 3)},
        'TBM_K_M_Noise_snr_1': {'model_type': 'TBM_conv1d', 'input_shape': (1024, 3)},
        'TBM_K_M_Noise_snr_0': {'model_type': 'TBM_conv1d', 'input_shape': (1024, 3)},
        'TBM_K_M_Noise_snr_-1': {'model_type': 'TBM_conv1d', 'input_shape': (1024, 3)},
        'TBM_K_M_Noise_snr_-3': {'model_type': 'TBM_conv1d', 'input_shape': (1024, 3)},
        'TBM_K_M_Noise_snr_-5': {'model_type': 'TBM_conv1d', 'input_shape': (1024, 3)},
        'TBM_K_M_Noise_snr_-7': {'model_type': 'TBM_conv1d', 'input_shape': (1024, 3)},
        'TBM_K_M_Noise_snr_-10': {'model_type': 'TBM_conv1d', 'input_shape': (1024, 3)},
    }
    
    # 如果指定了模型变体，直接使用
    if model_variant:
        if dataset_name in tbm_datasets:
            config = tbm_datasets[dataset_name].copy()
            config['model_type'] = model_variant
            return config
    
    # 合并配置
    all_configs = {**image_datasets, **tbm_datasets}
    
    if dataset_name in all_configs:
        return all_configs[dataset_name]
    else:
        # 默认使用图像模型
        print(f"警告: 未找到数据集 {dataset_name} 的配置，使用默认图像模型配置")
        return {'model_type': 'Q_Net_image', 'input_shape': (1, 28, 28)}


def main():

    set_random_seed(42)  # 设置全局随机数种子以确保可重复性
    
    # 创建TBM数据集配置列表，每个元素包含数据集名称和对应的rho值
    tbm_configs = [
        ('TBM_K_M_Noise', 0.01),  # 使用统一的数据集名称，我们将在ImbalancedDataset类中读取指定文件
    ]
    
    # 定义要测试的奖励倍数列表
    reward_multipliers = [1]
    
    # 定义要测试的折扣因子列表
    discount_factors = [0.1]
    
    # 使用固定的模型变体
    model_variants = ['TBM_conv1d_1layer', 'TBM_conv1d', 'TBM_conv1d_3layer']
    
    # 使用绝对路径
    save_dir = '/workspace/RL/DQNimb/multi_class_results'
    
    # 创建保存目录（如果不存在）
    os.makedirs(save_dir, exist_ok=True)
    
    # 多重循环：先遍历模型变体，数据集配置，再遍历奖励倍数和折扣因子
    for model_variant in model_variants:
        for dataset_name, rho in tbm_configs:
            for reward_multiplier in reward_multipliers:
                for discount_factor in discount_factors:
                    print(f"\n{'='*70}") 
                    print(f"开始处理数据集: {dataset_name}, 不平衡率: {rho}, 模型变体: {model_variant}")
                    print(f"奖励倍数: {reward_multiplier}, 折扣因子: {discount_factor}")
                    print(f"{'='*70}")
                    
                    # 获取模型配置
                    model_config = get_model_config(dataset_name, model_variant)
                    model_type = model_config['model_type']
                    input_shape = model_config['input_shape']
                    
                    print(f"数据集: {dataset_name}")
                    print(f"选择的模型类型: {model_type}")
                    print(f"输入形状: {input_shape}")
                    print(f"奖励倍数: {reward_multiplier}")
                    print(f"折扣因子: {discount_factor}")
                    
                    try:
                        # 创建数据集
                        dataset = ImbalancedDataset(dataset_name=dataset_name, rho=rho, batch_size=64)
                            
                        # 直接获取训练和测试的dataloader
                        train_loader, test_loader = dataset.get_dataloaders()
                        
                        # 获取分类数量
                        num_classes = 9  # 固定为9个类别（0-8）
                        
                        if TEST_ONLY:
                            print("测试模式: 加载多个模型并分别评估")
                            
                            # 设置与训练时相同的参数
                            num_runs = 10
                            training_ratio = 1  # 使用与训练时相同的ratio
                            
                            # 根据模型类型创建相应的模型，指定输出维度为9
                            if model_type == 'TBM_conv1d':
                                q_net = TBM_conv1d(input_shape, output_dim=num_classes)
                            else:
                                raise ValueError(f"不支持的模型类型: {model_type}")
                            
                            device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
                            q_net.to(device)
                            
                            # 循环加载每次训练保存的模型并评估
                            for run in range(1, num_runs + 1):
                                # 生成模型文件名，包含奖励倍数和折扣因子
                                model_filename = f'{dataset_name}_rho{rho}_{model_type}_reward{reward_multiplier}_gamma{discount_factor}_训练完成比{training_ratio}_第{run}次.pth'
                                model_path = os.path.join(save_dir, model_filename)
                                
                                if os.path.exists(model_path):
                                    print(f"\n{'='*50}")
                                    print(f"加载并评估第 {run} 个模型: {model_filename}")
                                    print(f"{'='*50}")
                                    
                                    # 加载模型
                                    q_net.load_state_dict(torch.load(model_path), strict=False)
                                    print(f"成功加载模型: {model_path}")
                                    
                                    # 评估模型，传递奖励倍数和折扣因子
                                    evaluate_model(q_net, test_loader, save_dir=save_dir, 
                                                dataset_name=dataset_name, training_ratio=training_ratio, rho=rho, 
                                                dataset_obj=dataset, run_number=run, model_type=model_type,
                                                reward_multiplier=reward_multiplier, discount_factor=discount_factor)
                                else:
                                    print(f"警告: 未找到模型文件 {model_path}")
                        else:
                            print("训练模式: 将进行模型训练和评估")

                            # 运行10次训练
                            num_runs = 10
                            print(f"开始进行 {num_runs} 次训练，模型类型: {model_type}, 奖励倍数: {reward_multiplier}, 折扣因子: {discount_factor}")
                            for run in range(1, num_runs + 1):
                                
                                print(f"\n{'='*50}")
                                print(f"开始第 {run} 次训练")
                                print(f"使用模型类型: {model_type}, 奖励倍数: {reward_multiplier}, 折扣因子: {discount_factor}")
                                print(f"{'='*50}")
                                
                                # 每次创建新的分类器实例，传入模型类型、奖励倍数和折扣因子
                                classifier = MyRL(input_shape, rho=rho, model_type=model_type, 
                                                reward_multiplier=reward_multiplier, discount_factor=discount_factor,
                                                num_classes=num_classes)
                                
                                # 开始训练，直接使用数据集对象而不是dataloader
                                classifier.train(dataset)
                                

                                # 使用MyRL类中的ratio参数
                                training_ratio = classifier.ratio
                                
                                # 生成带数据集名称、不平衡率、模型类型、奖励倍数、折扣因子、训练完成比例和序号的模型文件名
                                model_filename = f'{dataset_name}_rho{rho}_{model_type}_reward{reward_multiplier}_gamma{discount_factor}_训练完成比{training_ratio}_第{run}次.pth'
                                numbered_model_path = os.path.join(save_dir, model_filename)
                                
                                # 保存模型
                                torch.save(classifier.q_net.state_dict(), numbered_model_path)
                                print(f"模型已保存到 {numbered_model_path}")
                                
                                # 评估模型，传递奖励倍数和折扣因子
                                evaluate_model(classifier.q_net, test_loader, save_dir=save_dir, 
                                            dataset_name=dataset_name, training_ratio=training_ratio, rho=rho, 
                                            dataset_obj=dataset, run_number=run, model_type=model_type,
                                            reward_multiplier=reward_multiplier, discount_factor=discount_factor)
                                

                                # 绘制并保存损失曲线
                                loss_filename = f'{dataset_name}_rho{rho}_{model_type}_reward{reward_multiplier}_gamma{discount_factor}_训练完成比{training_ratio}_第{run}次_loss.png'
                                loss_path = os.path.join(save_dir, loss_filename)
                                classifier.plot_loss(loss_path)
                                
                                # 保存loss_history数据
                                loss_history_filename = f'{dataset_name}_rho{rho}_{model_type}_reward{reward_multiplier}_gamma{discount_factor}_训练完成比{training_ratio}_第{run}次_loss_history.npy'
                                loss_history_path = os.path.join(save_dir, loss_history_filename)
                                np.save(loss_history_path, np.array(classifier.loss_history))
                                print(f"损失历史数据已保存到 {loss_history_path}")
                                
                                print(f"第 {run} 次训练完成")
                    
                    except Exception as e:
                        print(f"处理配置 {dataset_name}, rho={rho}, reward={reward_multiplier}, gamma={discount_factor} 时出错: {e}")
                        import traceback
                        traceback.print_exc()
                        continue  # 继续下一个配置

if __name__ == "__main__":
    main()

# python /workspace/RL/DQNimb/main.py









